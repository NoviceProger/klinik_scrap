import time, random
from builtins import print

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.options import Options
import time
# import pyautogui	#pip install PyAutoGUI
from webdriver_manager.chrome import ChromeDriverManager

## всі клініки
url = "https://www.barmer-kliniksuche.de/suche/suchergebnis.php?searchdata%5Bplzort%5D=Berlin&searchdata%5Blocation%5D=6701%7CAGS&searchdata%5Bmaxdistance%5D=bundesweit&searchcontrol%5Border%5D=distance&searchcontrol%5Blimit_offset%5D=50&searchcontrol%5Blimit_num%5D=50"
options = Options()
# user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36'
# options.add_argument('user-agent={0}'.format(user_agent))
# accept_language = "en-US,en;q=0.9"
# options.add_argument('accept-language={0}'.format(accept_language))
options.add_argument("user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.163 Safari/537.36")
# options = webdriver.ChromeOptions()
# driver = webdriver.Chrome(executable_path="D:/My_PyCharm/Programs for python/Chromedriver/chromedriver.exe",
# 						  options=options)
# driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
# wait = WebDriverWait(driver, 20)
# # action = ActionChains(driver)
# hhh = driver.get(url)
# time.sleep(5)
# gear_xp = '//*[@id="more_btn"]'
# start_time = time.time()
# k = 1
# while k < 55:
# 	try:
# 		WebDriverWait(driver, 40).until(EC.element_to_be_clickable((By.XPATH, gear_xp)))
# 		driver.find_element(By.XPATH, gear_xp).click()
# 		print(f"True. k={k}")
# 	except:
# 		print(f"False. k={k}")
# 		k = 55
# 		pass
# 	k += 1
#
# finish_time = time.time() - start_time
# print(f'GEAR: {finish_time}')
# source_html = driver.page_source
# # source_html = driver.page_source
# #//*[@id="more_btn"]
# # //*[@id="more_btn"]
# # запис в ХТМЛ-файл
# with open(f"index1.html", 'w', encoding='utf-8') as file:
# 	file.write(source_html)
# # for i in range(4):
# # 	# pyautogui.moveRel(100, 0, duration=0.25)
# # 	pyautogui.moveRel(0, 100, duration=0.25)
# # 	# pyautogui.moveRel(-100, 0, duration=0.25)
# # 	pyautogui.moveRel(0, -100, duration=0.25)
# # find_button = '//*[@id="maincontent"]/main/div/div[3]/div[3]/nav/ul/li[3]/button'
# # driver.find_element(By.XPATH, find_button).click()
# # time.sleep(random.randint(2, 5))
#
# driver.close()

## по індексу(PLZ) з файлу XLSX
import requests
from openpyxl import Workbook
from openpyxl import load_workbook

# 'https://www.barmer-kliniksuche.de/suche/suchergebnis.php?searchcontrol%5Border%5D=distance&searchdata%5Bautocomplete%5D=&searchdata%5Bplzort%5D=52074&searchdata%5Blocation%5D=&searchdata%5Bmaxdistance%5D=0'
# 'https://www.barmer-kliniksuche.de/suche/suchergebnis.php?searchcontrol%5Border%5D=distance&searchdata%5Bautocomplete%5D=&searchdata%5Bplzort%5D=10115&searchdata%5Blocation%5D=&searchdata%5Bmaxdistance%5D=0'
wb = load_workbook(filename='zuordnung_plz_ort.xlsx')
sheet = wb['Worksheet']
row_count = sheet.max_row
list_search = []
for row in range(2, 5):
	str1 = sheet[f"C{row}"].value
	plz = str(str1).split()
	url = f"https://www.barmer-kliniksuche.de/suche/suchergebnis.php?searchcontrol%5Border%5D=distance&searchdata%5Bautocomplete%5D=&searchdata%5Bplzort%5D={plz}&searchdata%5Blocation%5D=&searchdata%5Bmaxdistance%5D=0"
	root = url
	result_get = requests.get(root)
	content = result_get.text
	url_fin = soup.find('tbody', id='search_result_table_body')
	# url_next = url_fin.find_all('tr')
	# запис в ХТМЛ-файл
	with open(f"HTML/index_{plz}.html", 'w', encoding='utf-8') as file:
		file.write(content)


# from bs4 import BeautifulSoup
# import json
#
# result = []
# links_result = []
# HTMLFile = open(f"index1.html", "r", encoding='utf-8')
# # Reading the file
# index = HTMLFile.read()
# soup = BeautifulSoup(index, 'lxml')
# # 5 пошук потрібних значень
# url_fin = soup.find('tbody', id='search_result_table_body')
# url_next = url_fin.find_all('tr')
# # print(url_next)
# for item_list in url_next:
# 	# items_icons = item_list.find('small', class_='hospital-icons')
# 	items = item_list.find_all('div')
# 	# print(items)
# 	policy1 = items[0].text.strip()
# 	find_url = items[1].find('a').get('href')
# 	policy2 = items[1].text.strip()
# 	policy3 = items[2].text.strip()
# 	links_result.append(
# 		{
# 			'url': find_url,
# 			'num': policy1,
# 			'name': policy2,
# 			'adress': policy3
# 		}
# 	)
#
# import requests
# import lxml, re
#
# num_klinik = 0
# for item_link in links_result:
# 	num_klinik += 1
# 	root = item_link['url']
# 	result_get = requests.get(root)
# 	content = result_get.text
# 	soup = BeautifulSoup(content, 'lxml')
# 	url_fin = soup.find('div', class_="col-sm-12")
# 	name = url_fin.find('strong').text.strip()
# 	# print(name)
# 	dom = lxml.etree.HTML(str(soup))
# 	adress = dom.xpath('//*[@id="maincontent"]/div[2]/div/div/text()[2]')[0].strip()
# 	# print(name1[0].strip())
# 	location = str(dom.xpath('//*[@id="maincontent"]/div[2]/div/div/text()[3]')[0].strip()).split(' ')
# 	telefon = dom.xpath('//*[@id="maincontent"]/div[2]/div/div/text()[4]')[0].strip()
# 	try:
# 		url_hospital_text = url_fin.find('a', class_="external").text
# 		# match = re.search(r'url=\S', url_hospital_text)
# 		if url_hospital_text and url_hospital_text.strip():
# 			url_hospital = url_hospital_text
# 		else:
# 			url_hospital1 = url_fin.find('a', class_="external").get('href')
# 			if str(url_hospital1).startswith('http'):
# 				url_hospital = url_hospital1
# 			else:
# 				url_hospital = f"https://www.barmer-kliniksuche.de{url_hospital1}"
# 	except:
# 		url_hospital = ""
#
# 	url_map1 = url_fin.find('p').find('a', class_="btn btn-primary btn-sm").get('href')
# 	if str(url_map1).startswith('http'):
# 		url_map = url_map1
# 	else:
# 		url_map = f"https://www.barmer-kliniksuche.de{url_map1}"
# 	date_update = str(dom.xpath('//*[@id="maincontent"]/div[7]/div[2]')[0].text.strip())
# 	# print(date_update)
#
# 	root = item_link['url'].replace('uebersicht', 'fachabteilungen')
# 	result_get = requests.get(root)
# 	content = result_get.text
# 	soup = BeautifulSoup(content, 'lxml')
# 	items_klinika = soup.find_all('div', class_='row rowtable-data')
#
# 	department = []
#
# 	for item_klinika in items_klinika:
# 		dep_url = item_klinika.find('div', class_='col-sm-8 col-xs-12 rowtable-title hyphenate').find('a').get('href')
# 		if str(dep_url).startswith('http'):
# 			department_url = dep_url
# 		else:
# 			department_url = f"https://www.barmer-kliniksuche.de{dep_url}"
# 		# print(department_url)
# 		root_dep = department_url
# 		result_get = requests.get(root_dep)
# 		content_dep = result_get.text
# 		soup_dep = BeautifulSoup(content_dep, 'lxml')
# 		name_dep = soup_dep.find('div', class_='title_hospital').find('h2', class_='hospital_specialist_department').text.strip()
# 		dir_dep = soup_dep.find('div', class_='col-sm-7 col-xs-12').text.strip()
# 		cases_dep = soup_dep.find('div', class_='col-sm-3 col-xs-12 rowtable-title hyphenate').text
# 		# print(cases_dep)
# 		root_dep = department_url.replace('uebersicht', 'diagnosen-krankheiten-icd')
# 		result_get = requests.get(root_dep)
# 		content_dep = result_get.text
# 		soup_dep = BeautifulSoup(content_dep, 'lxml')
# 		diagnoses = []
# 		items_diagnoses = soup_dep.find_all('div', class_='row rowtable-data')
# 		for item_diagnos in items_diagnoses:
# 			# print(item_diagnos)
# 			num_diagnosis = item_diagnos.find('div', class_='col-sm-1 col-xs-12 rowtable-title').text.strip()
# 			# print(num_diagnosis)
# 			diagnosis = item_diagnos.find('div',
# 						class_='col-sm-4 col-xs-12 rowtable-title hyphenate').text.strip().split('(')[0].strip()
# 			amount = item_diagnos.find('div', class_='col-sm-1 col-xs-8 rowtable-title').text.strip()
# 			# //*[@id="maincontent"]/div[2]/div[37]/div[2]/text() #//*[@id="maincontent"]/div[2]/div[38]/div[2]/text()
# 			diagnoses.append(
# 				{
# 					'num_diagnosis': num_diagnosis,
# 					'diagnosis': diagnosis,
# 					'amount': amount
# 				}
# 			)
# 			if int(num_diagnosis[:-1]) > 5:
# 				break
# 		root_pers = department_url.replace('uebersicht', 'personal')
# 		result_get = requests.get(root_pers)
# 		content_pers = result_get.text
# 		soup_dep = BeautifulSoup(content_pers, 'lxml')
# 		personal = []
# 		items_personal = soup_dep.find_all('div', class_='row rowtable-data')
# 		num = 0
# 		for item_personal in items_personal:
# 			# print(item_personal)
# 			num += 1
# 			if num > 5:
# 				break
# 			name_personal = item_personal.find('div',
# 										  class_='col-sm-6 col-xs-12 rowtable-title hyphenate').text.strip()
# 			amount_personal = item_personal.find('div', class_='col-sm-2 col-xs-12 rowtable-title hyphenate').text.strip()
# 			personal.append(
# 				{
# 					'name_personal': name_personal,
# 					'amount': amount_personal,
# 					'stationar': str(name_personal).find('stationär') != -1
# 				}
# 			)
# 		department.append(
# 			{
# 				'name': name_dep,
# 				'boss': dir_dep,
# 				'cases': cases_dep,
# 				'diagnoses': diagnoses,
# 				'personal': personal
# 			}
# 		)
# 		# print(department)
# 	result.append(
# 		{
# 			# 'type': url_item['type'],
# 			'name': name,
# 			'adress': adress,
# 			'plz': location[0],
# 			'ort': location[1],
# 			'telefon': telefon,
# 			'website': url_hospital,
# 			'google_map': url_map,
# 			'date_update': date_update,
# 			'departments': department
# 		}
# 	)
# 	print(f"Опрацьовано {num_klinik} з {len(links_result)}")
#
# # 	6 запис у файл JSON
# with open(f"result.json", 'w', encoding='utf-8') as file:
# 	json.dump(result, file, indent=4, ensure_ascii=False)